import json

from xadmin.util import vendor, xstatic
from xadmin.widgets import AdminFileWidget
import django.forms as django_forms
from django.conf import settings
from openpyxl.reader.excel import load_workbook
import collections
from irpf.views.base import AdminFormView


class XlsxViewerForm(django_forms.Form):
	filestream = django_forms.FileField(label="Arquivo de dados Excel (xlsx)",
	                                    widget=AdminFileWidget)


class AdminXlsxViewer(AdminFormView):
	"""Visualizador de dados de um arquivo Excel (xlsx)"""
	template_name = "irpf/adminx_report_irpf_viewer.html"
	form_class = XlsxViewerForm
	form_method_post = True

	title = "Visualizador de arquivos Excel"

	def process_sheet(self, ws):
		if settings.DEBUG:
			print("SHEET ", ws.title)
		data = collections.OrderedDict()

		rows = ws.iter_rows()

		headers = []
		for cell in next(rows):
			headers.append(cell.value)

		data['title'] = ws.title
		data['headers'] = headers
		items = []

		if settings.DEBUG:
			print(" / ".join(headers))

		for row in rows:
			cells = []
			for index, cell in enumerate(row):
				cells.append(str(cell.value))

			items.append(cells)
			if settings.DEBUG:
				print(" / ".join(cells))

		data['items'] = json.dumps(items)
		return data

	def file_handle(self, file):
		wb = None
		sheets = []
		with file:
			try:
				wb = load_workbook(filename=file, data_only=True)
				for sheet_name in wb.sheetnames:
					sheet_item = self.process_sheet(wb[sheet_name])
					sheets.append(sheet_item)
			finally:
				if wb:
					wb.close()
		return sheets

	def get_context(self):
		context = super().get_context()
		context['dt_language_url'] = xstatic('datatables.lang')[0]
		return context

	def get_media(self):
		media = super().get_media()
		media += vendor("datatables.js", "datatables.css")
		media += django_forms.Media(js=[
			'irpf/js/irpf.xlsx.viewer.js'
		], css={
			'screen': ("irpf/css/irpf.xlsx.viewer.css",)
		})
		return media

	def get_success_url(self):
		return self.get_admin_url("xlsx_viewer")

	def form_valid(self, form):
		filestream = form.cleaned_data["filestream"]
		sheets = self.file_handle(filestream.file)

		context = self.get_context_data()
		context['sheets'] = sheets

		return self.render_to_response(context)

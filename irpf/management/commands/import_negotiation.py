import argparse

from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from irpf.models import Negotiation


class Command(BaseCommand):
	help = """imports data from the xlsx file with information on the year's negotiations."""

	storage_model = Negotiation
	storage_ops = storage_model._meta

	def add_arguments(self, parser):
		parser.add_argument("--filepath", type=argparse.FileType('rb'), required=True)

	def get_fields_map(self):
		fields = {}
		for field in self.storage_ops.get_fields():
			if hasattr(field, "sheet_header"):
				fields[field.sheet_header] = field
		return fields

	def process_sheet(self, ws):
		print("SHEET ", ws.title)

		rows = ws.iter_rows()

		headers = []
		for cell in next(rows):
			headers.append(cell.value)

		print(" / ".join(headers))

		fields = self.get_fields_map()
		data = {}
		for row in rows:
			cells = []
			for index, cell in enumerate(row):
				try:
					field_name = fields[headers[index]].name
				except KeyError:
					continue
				data[field_name] = cell.value
				cells.append(str(cell.value))
			print(" / ".join(cells))
			self.storage_model.objects.get_or_create(**data)
			print(data)

	def handle(self, *args, **options):
		wb = None
		with options['filepath'] as filepath:
			try:
				wb = load_workbook(
					filename=filepath,
					data_only=True
				)

				for sheet_name in wb.sheetnames:
					self.process_sheet(wb[sheet_name])
			finally:
				if wb:
					wb.close()


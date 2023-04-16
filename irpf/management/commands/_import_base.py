import argparse

from django.contrib.auth import get_user_model, get_permission_codename
from django.core.exceptions import ValidationError
from django.core.management.base import BaseCommand
from django.db.transaction import atomic
from guardian.shortcuts import assign_perm
from openpyxl import load_workbook

User = get_user_model()


class UserType:
	def __init__(self, queryset):
		self.queryset = queryset
		self.model = queryset.model
		self.opts = self.model._meta

	def __call__(self, value):
		try:
			pk = self.opts.pk.to_python(value)
		except ValidationError:
			raise argparse.ArgumentTypeError(f"enter a valid primary key.")
		try:
			return self.queryset.get(pk=pk)
		except self.model.DoesNotExist:
			raise argparse.ArgumentTypeError(f"user not found!")


class Command(BaseCommand):
	help = """imports data from the xlsx file with information on the earnings."""
	model_permissions = ["view", "add", "change", "delete"]

	storage_model = None
	storage_ops = None

	def add_arguments(self, parser):
		parser.add_argument("--filepath", type=argparse.FileType('rb'), required=True)
		parser.add_argument("--user", type=UserType(User.objects.filter(is_active=True)),
		                    required=True)

	def get_fields_map(self):
		fields = {}
		for field in self.storage_ops.get_fields():
			if hasattr(field, "sheet_header"):
				fields.setdefault(field.sheet_header, []).append(field)
		return fields

	def _assign_perm(self, instance, user=None):
		"""Adiciona permissões padrão para a instância x usuário"""
		opts = instance._meta
		if user is None:
			user = instance.user
		for name in self.model_permissions:
			permission_codename = get_permission_codename(name, opts)
			assign_perm(permission_codename, user, instance)

	@atomic
	def save_instance(self, **data):
		if hasattr(self.storage_model, "import_before_save_data"):
			data = self.storage_model.import_before_save_data(**data)
		instance, created = self.storage_model.objects.get_or_create(**data)
		if created:
			self._assign_perm(instance)

	def process_sheet(self, ws, options):
		verbosity, level = options.get('verbosity', 0), 2
		if verbosity > level:
			print("SHEET ", ws.title)

		rows = ws.iter_rows()

		headers = []
		for cell in next(rows):
			headers.append(cell.value)

		if verbosity > level:
			print(" / ".join(headers))

		fields = self.get_fields_map()
		user = options['user']
		for row in rows:
			cells = []
			data = {'user': user}
			for index, cell in enumerate(row):
				try:
					header_fields = fields[headers[index]]
				except KeyError:
					continue
				for field in header_fields:
					data[field.name] = cell.value
				cells.append(str(cell.value))
			if verbosity > level:
				print(" / ".join(cells))
			self.save_instance(**data)

	def handle(self, *args, **options):
		wb = None
		with options['filepath'] as filepath:
			try:
				wb = load_workbook(
					filename=filepath,
					data_only=True
				)

				for sheet_name in wb.sheetnames:
					self.process_sheet(wb[sheet_name], options)
			finally:
				if wb:
					wb.close()
